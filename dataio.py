import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def save_hl_sheet(filepath, dataframe, hl_locations):
    dataframe.to_excel(filepath, index=False, engine="openpyxl")
    wb = load_workbook(filepath)
    ws = wb.active
    # Define yellow fill style
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    # Apply yellow fill to empty cells
    for row, col in hl_locations:
        ws.cell(row=row + 2, column=col + 1).fill = yellow_fill

    # Save the modified file
    wb.save(filepath)


def write_data(out_data, filename):
    with open(filename, "w") as file:
        json.dump(out_data, file, indent=4)


def write_meta_data(out_data):
    write_data(out_data, "metadata.json")


# this will be much faster than performing operations over the API access
def sheet_to_json(sheet):
    json_rows = sheet.get_all_records()
    write_data(json_rows, "sheet-raw.json")


def read_sheet_data():
    return read_data("sheet-raw.json")


def read_meta_data():
    return read_data("metadata.json")


def read_data(filename):
    with open(filename, "r") as file:
        data = json.load(file)
    return data
